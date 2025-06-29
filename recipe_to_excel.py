import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from coze_api_utils import CozeAPIClient
import json

def save_to_excel(recipe_data, excel_path, current_id):
    """
    将菜谱数据保存到Excel表格
    
    参数:
        recipe_data: 菜谱数据字典
        excel_path: Excel文件路径
        current_id: 当前菜谱的ID
    """
    try:
        # 获取当前时间
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 从recipe_data中获取数据
        name = recipe_data.get('name', '')
        cook_time = recipe_data.get('cook_time', '')
        calories = recipe_data.get('calories', '')
        image = recipe_data.get('image', '')
        description = recipe_data.get('description', '')
        # 将列表和字典类型的字段转换为JSON字符串
        steps = json.dumps(recipe_data.get('steps', []), ensure_ascii=False)
        tools = json.dumps(recipe_data.get('tools', []), ensure_ascii=False)
        prep_steps = json.dumps(recipe_data.get('prep_steps', []), ensure_ascii=False)
        tips = json.dumps(recipe_data.get('tips', []), ensure_ascii=False)
        difficulty = recipe_data.get('difficulty', '')
        
        # 创建新行数据
        new_row = {
            'id': current_id,
            'name': name,
            'cook_time': cook_time,
            'calories': calories,
            'image': image,
            'description': description,
            'steps': steps,
            'tools': tools,
            'prep_steps': prep_steps,
            'tips': tips,
            'difficulty': difficulty,
            'created_at': current_time,
            'updated_at': current_time
        }
        
        # 加载Excel文件（保留所有工作表）
        book = load_workbook(excel_path)
        
        # 检查是否存在recipes工作表
        if 'recipes' not in book.sheetnames:
            # 如果不存在，创建新的工作表
            sheet = book.create_sheet('recipes')
            # 添加表头 - 更新表头，添加新字段
            sheet.append(['id', 'name', 'cook_time', 'calories', 'image', 'description', 'steps', 'tools', 'prep_steps', 'tips', 'difficulty', 'created_at', 'updated_at'])
        else:
            # 如果存在，获取该工作表
            sheet = book['recipes']
            
            # 检查是否需要更新表头（添加新字段）
            header_row = [cell.value for cell in sheet[1]]
            required_headers = ['id', 'name', 'cook_time', 'calories', 'image', 'description', 'steps', 'tools', 'prep_steps', 'tips', 'difficulty', 'created_at', 'updated_at']
            
            # 检查是否缺少某些表头
            missing_headers = [header for header in required_headers if header not in header_row]
            if missing_headers:
                print(f"表头缺少以下字段：{', '.join(missing_headers)}")
                print("将在现有表头后添加缺失字段")
                
                # 添加缺失的列头
                for header in missing_headers:
                    sheet.cell(row=1, column=len(header_row) + 1).value = header
                    header_row.append(header)
        
        # 添加新行
        # 获取当前表头
        current_headers = [cell.value for cell in sheet[1]]
        
        # 准备新行数据，按表头顺序
        row_data = []
        for header in current_headers:
            if header in new_row:
                row_data.append(new_row[header])
            else:
                row_data.append('')  # 对于未知字段，添加空值
                
        # 添加新行
        sheet.append(row_data)
        
        # 保存Excel文件
        book.save(excel_path)
        print(f"成功将菜谱 '{name}' 保存到Excel")
        return True
    except Exception as e:
        print(f"保存到Excel时出错: {e}")
        return False

def get_last_id(excel_path):
    """
    获取Excel中最后一个ID
    
    参数:
        excel_path: Excel文件路径
    
    返回:
        最后一个ID，如果没有则返回0
    """
    try:
        # 检查文件是否存在
        if not os.path.exists(excel_path):
            return 0
        
        # 加载Excel文件
        book = load_workbook(excel_path)
        
        # 检查是否存在recipes工作表
        if 'recipes' not in book.sheetnames:
            return 0
        
        # 获取工作表
        sheet = book['recipes']
        
        # 获取行数
        row_count = sheet.max_row
        
        # 如果只有表头或为空，返回0
        if row_count <= 1:
            return 0
        
        # 获取最后一行的ID
        last_id = sheet.cell(row=row_count, column=1).value
        
        # 如果最后一行的ID为空，从倒数第二行开始查找非空ID
        if last_id is None:
            for row in range(row_count-1, 0, -1):
                last_id = sheet.cell(row=row, column=1).value
                if last_id is not None:
                    break
        
        # 如果找不到有效ID，返回0
        return int(last_id) if last_id is not None else 0
    except Exception as e:
        print(f"获取最后ID时出错: {e}")
        return 0

def main():
    # 创建API客户端
    client = CozeAPIClient()
    
    # 检查是否能获取到令牌
    if not client.get_token(interactive=False):
        return
    
    # 文件路径
    recipes_file = "家常菜菜谱名称.txt"
    excel_path = "data/database.xlsx"
    
    # 从文件中读取菜谱名称
    with open(recipes_file, 'r', encoding='utf-8') as file:
        recipe_names = [line.strip() for line in file.readlines() if line.strip()]
    
    # 获取最后一个ID
    last_id = get_last_id(excel_path)
    print(f"Excel中的最后ID: {last_id}")
    
    # 为每个菜谱调用API并保存到Excel
    for index, recipe_name in enumerate(recipe_names, 1):
        print(f"\n{'-' * 50}")
        print(f"处理第{index}个菜谱: {recipe_name}")
        print(f"{'-' * 50}")
        
        # 调用API获取菜谱数据
        recipe_data = client.get_recipe(recipe_name, interactive=False, verbose=False)
        
        print(f"获取到的菜谱数据类型: {type(recipe_data)}")
        if recipe_data is None:
            print(f"获取菜谱 '{recipe_name}' 失败: 返回了None")
            continue
            
        if isinstance(recipe_data, str):
            print(f"获取菜谱 '{recipe_name}' 失败: 返回了字符串而非JSON对象")
            print(f"返回内容: {recipe_data}")
            continue
            
        if isinstance(recipe_data, dict):
            # 检查返回的字典是否包含必要的字段
            required_fields = ['name', 'steps']
            missing_fields = [field for field in required_fields if field not in recipe_data]
            
            if missing_fields:
                print(f"获取菜谱 '{recipe_name}' 返回的数据缺少必要字段: {', '.join(missing_fields)}")
                print(f"返回的字段有: {', '.join(recipe_data.keys())}")
                continue
                
            # 保存到Excel
            current_id = last_id + index
            try:
                success = save_to_excel(recipe_data, excel_path, current_id)
                if success:
                    print(f"成功保存菜谱 '{recipe_name}' 到Excel")
                else:
                    print(f"保存菜谱 '{recipe_name}' 到Excel失败")
            except Exception as e:
                print(f"保存菜谱 '{recipe_name}' 到Excel时发生异常: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"获取菜谱 '{recipe_name}' 失败: 返回了意外的数据类型 {type(recipe_data)}")
    
    print("\n所有菜谱处理完成")

if __name__ == "__main__":
    main() 